#!/usr/bin/python
# -*- coding: utf-8 -*-
# File Name: data.py
# Author: Changsheng Zhang
# mail: zhangcsxx@gmail.com


from openpyxl.reader.excel import load_workbook
import random
import time

def loadWord(vocab,meaning):
	fileName = "data.xlsx"
	wb = load_workbook(fileName)
	sheetnames = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheetnames[0])
	workNum = ws.get_highest_row()
	for rx in range(1,ws.get_highest_row()+1):
		tempVocab = ws.cell(row=rx,column =1).value
		tempMeaning =ws.cell(row=rx,column =2).value
		vocab.append(tempVocab)
		meaning.append(tempMeaning)
		#print tempMeaning
        



#def strategy(vocab,mean):


def seqRun(vocab,mean,newVocabIndex,begin):
        #loopNum = input("请输出要记忆的单词数量:")
	maxNum = len(vocab)
        beginIndex = 0
        endIndex = 0
        
        for i in range(maxNum):
            if vocab[i][0]==begin:
                beginIndex =i
                break
        for i in range(beginIndex,maxNum):
            if begin!='z':
                if vocab[i][0]!=begin:
                    endIndex =i
                    break
            else:
                endIndex = maxNum 
        loopNum = endIndex - beginIndex

        vocabIndex = beginIndex 
        for i in range(loopNum):

		print "\033[1;34;36m"
		print vocab[vocabIndex]
		print "\033[0m"

		print "您是否掌握这个单词？"
		getVocab = raw_input("是(q)/否(任意+回车):")
		if getVocab == 'y' or getVocab =='q':
			print "\033[1;31;34m"
			print mean[vocabIndex]
			print "\033[0m"

		else:
			print "\033[1;31;31m"
			print mean[vocabIndex]
			print "\033[0m"

			newVocabIndex.append(vocabIndex)

		vocabIndex = vocabIndex +1

def randomRun(vocab,mean,newVocabIndex):

	loopNum = input("请输出要记忆的单词数量:")
	maxNum = len(vocab)
	for i in range(loopNum):
		vocabIndex = random.randint(1,maxNum)-1
		print "\033[1;34;36m"
		print vocab[vocabIndex]
		print "\033[0m"


		print "您是否掌握这个单词？"

		getVocab = raw_input("是(q)/否(任意):")
		if getVocab == 'y' or getVocab =='Y':
			print "\033[1;31;34m"
			print mean[vocabIndex]
			print "\033[0m"

		else:
			print "\033[1;31;31m"
			print mean[vocabIndex]
			print "\033[0m"

			newVocabIndex.append(vocabIndex)



def saveToFile(vocab,mean,newVocabIndex):
	fileName = time.strftime('%Y-%m-%d-%H-%M',time.localtime(time.time()))+".md"
	f = open(fileName,"w")
	for i in range(len(newVocabIndex)):
		f.write("+ ")
		f.write(vocab[newVocabIndex[i]])
		f.write("  ")
		f.write(mean[newVocabIndex[i]].encode('utf-8'))
		f.write("\n")
	f.close()

def repeatVocab(vocab,mean,newVocabIndex):

	for i in range(len(newVocabIndex)):
		print "\033[1;34;36m"
		vocabIndex = newVocabIndex[i]
		print vocab[vocabIndex]
		print "\033[0m"

		print "您是否掌握这个单词？"
		getVocab = raw_input("是(q)/否(任意+回车):")
		if getVocab == 'y' or getVocab =='q':
			print "\033[1;31;34m"
			print mean[vocabIndex]
			print "\033[0m"

		else:
			print "\033[1;31;31m"
			print mean[vocabIndex]
			print "\033[0m"

			#newVocabIndex.append(vocabIndex)



if __name__ == '__main__':
	vocab  =[]
	mean = []
	newVocabIndex =[]

	loadWord(vocab,mean)
	print "请选择学习模式:"
	runMode =raw_input("随机(r)/顺序(s):")

	if runMode == 'r':
		randomRun(vocab,mean,newVocabIndex)
	else:
            begin = raw_input("请输入起始字母:")
	    seqRun(vocab,mean,newVocabIndex,begin)


#	print "重复刚才没有掌握的单词："
#	repeatVocab(vocab,mean,newVocabIndex)

	print "学习结束!没有掌握的单词已经保存在文件中。"
	saveToFile(vocab,mean,newVocabIndex)


